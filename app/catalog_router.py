import csv
import os
import threading
import unicodedata
from pathlib import Path

from fastapi import APIRouter, BackgroundTasks, HTTPException, Request, UploadFile
from fastapi.responses import JSONResponse

from app.database import get_db, DB_PATH
from app.sap_sync_worker import sync_status, start_async_sync

router = APIRouter(prefix="/api", tags=["Catalog"])

WAREHOUSES = ["01", "11", "15", "30"]

# Vista completa: usuarios autenticados
_PRODUCT_COLS = """
    p.sku,
    p.name,
    p.description,
    p.item_type,
    p.sell_item,
    p.price,
    p.images,
    p.image_url,
    p.woo_regular_price,
    p.woo_sale_price,
    COALESCE(s01.on_hand, 0) AS stock_01,
    COALESCE(s11.on_hand, 0) AS stock_11,
    COALESCE(s15.on_hand, 0) AS stock_15,
    COALESCE(s30.on_hand, 0) AS stock_30,
    (COALESCE(s01.on_hand, 0) + COALESCE(s11.on_hand, 0) +
     COALESCE(s15.on_hand, 0) + COALESCE(s30.on_hand, 0)) AS total_stock,
    p.location
"""

# Vista pública: stock tienda (B15) y web (B01+B11), precio SAP incluido solo como base de cálculo IVA
_PUBLIC_COLS = """
    p.sku,
    p.name,
    p.description,
    p.item_type,
    p.sell_item,
    p.images,
    p.image_url,
    p.price,
    p.woo_regular_price,
    p.woo_sale_price,
    COALESCE(s15.on_hand, 0) AS stock_tienda,
    (COALESCE(s01.on_hand, 0) + COALESCE(s11.on_hand, 0)) AS stock_web,
    (COALESCE(s01.on_hand, 0) + COALESCE(s11.on_hand, 0) +
     COALESCE(s15.on_hand, 0) + COALESCE(s30.on_hand, 0)) AS total_stock,
    p.location
"""

_JOINS = """
    FROM products p
    LEFT JOIN stock s01 ON p.sku = s01.sku AND s01.warehouse_code = '01'
    LEFT JOIN stock s11 ON p.sku = s11.sku AND s11.warehouse_code = '11'
    LEFT JOIN stock s15 ON p.sku = s15.sku AND s15.warehouse_code = '15'
    LEFT JOIN stock s30 ON p.sku = s30.sku AND s30.warehouse_code = '30'
"""


def _drive_url(file_id: str) -> str:
    return f"https://drive.google.com/thumbnail?id={file_id}&sz=w600"


def _expand_images(row: dict) -> dict:
    raw = row.get("images") or ""
    ids = [fid.strip() for fid in raw.split(",") if fid.strip()]
    urls = [_drive_url(fid) for fid in ids]
    # Fallback: si no hay IDs de Drive, usa la URL de WooCommerce
    if not urls and row.get("image_url"):
        urls = [row["image_url"]]
    row["image_urls"] = urls
    row["image_count"] = len(urls)
    return row


def _normalize(text: str) -> str:
    if not text:
        return ""
    return "".join(
        c for c in unicodedata.normalize("NFD", str(text)) if unicodedata.category(c) != "Mn"
    ).lower()


def _build_conditions(search: str, item_type: str, stock_filter: str, warehouses: str, sell_item: str, category: str, channel: str = "", is_auth: bool = True):
    conditions, args = [], []

    if search:
        search_terms = _normalize(search).split()
        for term in search_terms:
            like_term = f"%{term}%"
            conditions.append("(p.name_norm LIKE ? OR LOWER(p.sku) LIKE ? OR LOWER(p.description) LIKE ?)")
            args.extend([like_term, like_term, like_term])

    if item_type and item_type != "all":
        conditions.append("p.item_type = ?")
        args.append(item_type)

    # Auth: filtra por total 4 bodegas. Público: B01+B11 (web) + B15 (tienda)
    stock_expr = (
        "COALESCE(s01.on_hand,0)+COALESCE(s11.on_hand,0)+"
        "COALESCE(s15.on_hand,0)+COALESCE(s30.on_hand,0)"
        if is_auth else
        "COALESCE(s01.on_hand,0)+COALESCE(s11.on_hand,0)+COALESCE(s15.on_hand,0)"
    )
    if stock_filter == "instock":
        conditions.append(f"({stock_expr}) > 0")
    elif stock_filter == "outofstock":
        conditions.append(f"({stock_expr}) = 0")

    # Filtro de canal: tienda (B15), web (B01+B11), ambos
    if channel == "tienda":
        conditions.append("COALESCE(s15.on_hand, 0) > 0")
    elif channel == "web":
        conditions.append("(COALESCE(s01.on_hand, 0) + COALESCE(s11.on_hand, 0)) > 0")
    elif channel == "ambos":
        conditions.append("COALESCE(s15.on_hand, 0) > 0")
        conditions.append("(COALESCE(s01.on_hand, 0) + COALESCE(s11.on_hand, 0)) > 0")

    whs_alias = {"01": "s01", "11": "s11", "15": "s15", "30": "s30"}
    if warehouses:
        whs_list = [w.strip() for w in warehouses.split(",") if w.strip() in whs_alias]
        if whs_list:
            parts = " OR ".join(f"COALESCE({whs_alias[w]}.on_hand, 0) > 0" for w in whs_list)
            conditions.append(f"({parts})")

    if sell_item == "yes":
        conditions.append("p.sell_item = 1")

    if category:
        conditions.append("(',' || p.categories || ',') LIKE ?")
        args.append(f"%,{category},%")

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    return where, args


def migrate_csv_locations() -> None:
    """Importa ubicaciones del CSV al DB una sola vez (si la DB no tiene ubicaciones aún)."""
    csv_path = Path(__file__).parent.parent / "Ubicaciones Tienda SKU - Hoja 1.csv"
    if not csv_path.exists() or not os.path.exists(DB_PATH):
        return
    try:
        conn = get_db()
        count = conn.execute("SELECT COUNT(*) FROM products WHERE location != ''").fetchone()[0]
        conn.close()
        if count > 0:
            return
        locations: dict[str, str] = {}
        with open(csv_path, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                sku = (row.get("SKU") or "").strip().upper()
                if not sku:
                    continue
                locs = [
                    (row.get("UBICACION-1") or "").strip(),
                    (row.get("UBICACION-2") or "").strip(),
                    (row.get("UBICACION-3") or "").strip(),
                ]
                locs = [l for l in locs if l]
                if locs and sku not in locations:
                    locations[sku] = ", ".join(locs)
        if not locations:
            return
        conn = get_db()
        with conn:
            for sku, loc in locations.items():
                conn.execute(
                    "UPDATE products SET location=? WHERE UPPER(sku)=UPPER(?)",
                    (loc, sku),
                )
        conn.close()
    except Exception:
        pass


@router.get("/categories")
async def get_categories():
    if not os.path.exists(DB_PATH):
        return []
    conn = get_db()
    rows = conn.execute("SELECT slug, name FROM categories ORDER BY name").fetchall()
    conn.close()
    return [dict(r) for r in rows]


@router.get("/products")
async def get_products(
    request: Request,
    search: str = "",
    item_type: str = "all",
    stock_filter: str = "all",
    warehouses: str = "",
    sell_item: str = "all",
    category: str = "",
    channel: str = "",
    page: int = 1,
    page_size: int = 24,
):
    is_auth = bool(request.session.get("user"))
    cols    = _PRODUCT_COLS if is_auth else _PUBLIC_COLS

    if not os.path.exists(DB_PATH):
        return {
            "products": [],
            "authenticated": is_auth,
            "pagination": {"current_page": 1, "total_pages": 0, "total_items": 0, "page_size": page_size},
        }

    where, args = _build_conditions(search, item_type, stock_filter, warehouses, sell_item, category, channel, is_auth)

    conn = get_db()
    total_items = conn.execute(f"SELECT COUNT(*) {_JOINS} {where}", args).fetchone()[0]
    total_pages = max(1, (total_items + page_size - 1) // page_size)
    page = max(1, min(page, total_pages))
    offset = (page - 1) * page_size

    order_clause = "p.name"
    order_args: list = []
    if search:
        norm = _normalize(search.strip())
        order_clause = """
            CASE
                WHEN LOWER(p.sku) = LOWER(?)   THEN 0
                WHEN p.name_norm LIKE ?         THEN 1
                ELSE                                 2
            END, p.name
        """
        order_args = [search.strip(), f"%{norm}%"]

    rows = conn.execute(
        f"SELECT {cols} {_JOINS} {where} ORDER BY {order_clause} LIMIT ? OFFSET ?",
        args + order_args + [page_size, offset],
    ).fetchall()
    conn.close()

    return {
        "products": [_expand_images(dict(r)) for r in rows],
        "authenticated": is_auth,
        "pagination": {
            "current_page": page,
            "total_pages": total_pages,
            "total_items": total_items,
            "page_size": page_size,
        },
    }


@router.get("/product/{sku}")
async def get_product(request: Request, sku: str):
    is_auth = bool(request.session.get("user"))
    cols    = _PRODUCT_COLS if is_auth else _PUBLIC_COLS

    if not os.path.exists(DB_PATH):
        raise HTTPException(
            status_code=503,
            detail="Base de datos no inicializada. Ejecute una sincronización primero.",
        )

    conn = get_db()
    row = conn.execute(
        f"SELECT {cols} {_JOINS} WHERE UPPER(p.sku) = UPPER(?)",
        (sku.strip(),),
    ).fetchone()
    conn.close()

    if not row:
        raise HTTPException(status_code=404, detail=f"Producto con SKU '{sku}' no encontrado.")

    return {"status": "success", "authenticated": is_auth, "product": _expand_images(dict(row))}


@router.get("/sync-status")
async def get_sync_status():
    return sync_status


@router.post("/trigger-sync")
async def trigger_sync(background_tasks: BackgroundTasks):
    if sync_status["is_running"]:
        return {"message": "Sincronización ya en curso.", "status": sync_status}
    background_tasks.add_task(start_async_sync)
    return {"message": "Sincronización iniciada.", "status": sync_status}


@router.get("/sync-schedule")
async def get_sync_schedule():
    from app.main import scheduler, SYNC_INTERVAL_MINUTES
    job = scheduler.get_job("auto_sync")
    next_run = (
        job.next_run_time.strftime("%Y-%m-%d %H:%M:%S") if job and job.next_run_time else None
    )
    return {"interval_minutes": SYNC_INTERVAL_MINUTES, "next_sync": next_run}


@router.patch("/product/{sku}/location")
async def update_product_location(request: Request, sku: str, body: dict):
    _require_auth(request)
    location = (body.get("location") or "").strip()
    conn = get_db()
    row = conn.execute(
        "SELECT sku FROM products WHERE UPPER(sku)=UPPER(?)", (sku.strip(),)
    ).fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail=f"SKU '{sku}' no encontrado")
    conn = get_db()
    with conn:
        conn.execute(
            "UPDATE products SET location=? WHERE UPPER(sku)=UPPER(?)",
            (location, sku.strip()),
        )
    conn.close()
    return {"sku": sku, "location": location}


@router.post("/locations/bulk")
async def bulk_update_locations(request: Request, file: UploadFile):
    _require_auth(request)
    content = await file.read()
    filename = (file.filename or "").lower()
    locations: dict[str, str] = {}
    try:
        if filename.endswith((".xlsx", ".xls")):
            from io import BytesIO
            import openpyxl
            wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                d = dict(zip(headers, row))
                sku = str(d.get("SKU") or "").strip().upper()
                if not sku:
                    continue
                locs = [
                    str(d.get("UBICACION-1") or "").strip(),
                    str(d.get("UBICACION-2") or "").strip(),
                    str(d.get("UBICACION-3") or "").strip(),
                ]
                locs = [l for l in locs if l and l.lower() != "none"]
                if sku not in locations:
                    locations[sku] = ", ".join(locs)
        else:
            from io import StringIO
            text = content.decode("utf-8-sig")
            for row in csv.DictReader(StringIO(text)):
                sku = (row.get("SKU") or "").strip().upper()
                if not sku:
                    continue
                locs = [
                    (row.get("UBICACION-1") or "").strip(),
                    (row.get("UBICACION-2") or "").strip(),
                    (row.get("UBICACION-3") or "").strip(),
                ]
                locs = [l for l in locs if l]
                if sku not in locations:
                    locations[sku] = ", ".join(locs)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al leer archivo: {e}")

    if not locations:
        raise HTTPException(status_code=400, detail="No se encontraron datos válidos en el archivo.")

    conn = get_db()
    updated = 0
    with conn:
        for sku, loc in locations.items():
            result = conn.execute(
                "UPDATE products SET location=? WHERE UPPER(sku)=UPPER(?)",
                (loc, sku),
            )
            updated += result.rowcount
    conn.close()
    return {"updated": updated, "total": len(locations)}


# ── Upload de imágenes ─────────────────────────────────────────────────────────

upload_status = {
    "is_running": False,
    "done": 0,
    "total": 0,
    "errors": [],
    "message": "",
}


def _require_auth(request: Request):
    if not request.session.get("user"):
        raise HTTPException(status_code=401, detail="Autenticación requerida")


def _get_folder_id() -> str:
    fid = os.getenv("GDRIVE_FOLDER_ID", "")
    if not fid:
        raise HTTPException(status_code=500, detail="GDRIVE_FOLDER_ID no configurado")
    return fid


def _ext(filename: str, content_type: str = "") -> str:
    suffix = Path(filename).suffix.lower()
    if suffix in (".jpg", ".jpeg", ".png", ".webp", ".gif"):
        return ".jpg" if suffix == ".jpeg" else suffix
    ct_map = {"image/jpeg": ".jpg", "image/png": ".png", "image/webp": ".webp", "image/gif": ".gif"}
    return ct_map.get(content_type.split(";")[0].strip(), ".jpg")


@router.get("/upload-status")
async def get_upload_status():
    return upload_status


@router.post("/product/{sku}/images")
async def upload_product_images(request: Request, sku: str, files: list[UploadFile]):
    _require_auth(request)
    folder_id = _get_folder_id()
    sku = sku.strip().upper()

    conn = get_db()
    row = conn.execute("SELECT sku, images FROM products WHERE UPPER(sku)=UPPER(?)", (sku,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail=f"SKU '{sku}' no encontrado")

    existing_ids = [fid for fid in (row["images"] or "").split(",") if fid.strip()]
    next_idx = len(existing_ids) + 1

    from app.gdrive_client import get_service, upload_image, ids_to_urls
    svc = get_service()
    new_ids = list(existing_ids)

    for i, file in enumerate(files):
        data = await file.read()
        ext = _ext(file.filename or "", file.content_type or "")
        mime = (file.content_type or "image/jpeg").split(";")[0].strip()
        filename = f"{sku}-{next_idx + i}{ext}"
        file_id = upload_image(svc, folder_id, filename, data, mime)
        new_ids.append(file_id)

    conn = get_db()
    with conn:
        conn.execute("UPDATE products SET images=? WHERE UPPER(sku)=UPPER(?)", (",".join(new_ids), sku))
    conn.close()

    return {"sku": sku, "image_urls": ids_to_urls(",".join(new_ids)), "image_count": len(new_ids)}


@router.delete("/product/{sku}/images/{index}")
async def delete_product_image(request: Request, sku: str, index: int):
    """Elimina la imagen en la posición index (1-based) del producto."""
    _require_auth(request)
    sku = sku.strip().upper()

    conn = get_db()
    row = conn.execute("SELECT images FROM products WHERE UPPER(sku)=UPPER(?)", (sku,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail=f"SKU '{sku}' no encontrado")

    ids = [fid for fid in (row["images"] or "").split(",") if fid.strip()]
    if index < 1 or index > len(ids):
        raise HTTPException(status_code=400, detail=f"Índice {index} fuera de rango")

    file_id = ids.pop(index - 1)

    # Intentar borrar de Drive (si ya fue borrado manualmente, no es error)
    try:
        from app.gdrive_client import get_service, delete_file
        svc = get_service()
        delete_file(svc, file_id)
    except Exception:
        pass

    new_images = ",".join(ids)
    conn = get_db()
    with conn:
        conn.execute("UPDATE products SET images=? WHERE UPPER(sku)=UPPER(?)", (new_images, sku))
    conn.close()

    return {"sku": sku, "image_urls": [_drive_url(fid) for fid in ids], "image_count": len(ids)}


@router.post("/images/bulk-upload/check")
async def bulk_upload_check(request: Request, files: list[UploadFile]):
    """Recibe archivos, devuelve lista de conflictos sin subir nada."""
    _require_auth(request)

    parsed, parse_errors = [], []
    for file in files:
        stem = Path(file.filename or "").stem
        parts = stem.rsplit("-", 1)
        if len(parts) != 2 or not parts[1].isdigit():
            parse_errors.append(f"{file.filename}: nombre inválido (esperado SKU-N.ext)")
            continue
        parsed.append({"sku": parts[0].upper(), "idx": int(parts[1]), "name": file.filename})

    if parse_errors:
        raise HTTPException(status_code=400, detail={"parse_errors": parse_errors})

    skus = list({p["sku"] for p in parsed})
    conn = get_db()
    rows = conn.execute(
        f"SELECT sku, images FROM products WHERE UPPER(sku) IN ({','.join('?'*len(skus))})",
        skus,
    ).fetchall()
    conn.close()

    images_map = {
        r["sku"].upper(): len([x for x in (r["images"] or "").split(",") if x.strip()])
        for r in rows
    }
    not_found = [p["sku"] for p in parsed if p["sku"] not in images_map]
    conflicts = [p["name"] for p in parsed if p["idx"] <= images_map.get(p["sku"], 0)]

    return {"conflicts": conflicts, "not_found": not_found, "total": len(parsed)}


def _run_bulk_upload(file_data: list, folder_id: str):
    from app.gdrive_client import get_service, upload_image, delete_file

    upload_status.update({"is_running": True, "done": 0, "total": len(file_data), "errors": [], "message": "Subiendo..."})

    try:
        svc = get_service()
        conn = get_db()

        sku_ids: dict[str, list] = {}
        skus = list({f["sku"] for f in file_data})
        rows = conn.execute(
            f"SELECT sku, images FROM products WHERE UPPER(sku) IN ({','.join('?'*len(skus))})", skus
        ).fetchall()
        for r in rows:
            sku_ids[r["sku"].upper()] = [x for x in (r["images"] or "").split(",") if x.strip()]

        for item in file_data:
            sku, idx, data, mime, filename = item["sku"], item["idx"], item["data"], item["mime"], item["filename"]
            ids = sku_ids.get(sku, [])

            # Sobreescribir: eliminar ID anterior en esa posición
            if idx <= len(ids):
                delete_file(svc, ids[idx - 1])
                ids[idx - 1] = ""  # placeholder

            # Extender lista si el índice es mayor
            while len(ids) < idx:
                ids.append("")

            try:
                ext = _ext(filename, mime)
                file_id = upload_image(svc, folder_id, f"{sku}-{idx}{ext}", data, mime)
                ids[idx - 1] = file_id
                sku_ids[sku] = ids
            except Exception as e:
                upload_status["errors"].append(f"{sku}-{idx}: {e}")

            upload_status["done"] += 1

        # Guardar todos los cambios en DB
        with conn:
            for sku, ids in sku_ids.items():
                clean = ",".join(x for x in ids if x)
                conn.execute("UPDATE products SET images=? WHERE UPPER(sku)=UPPER(?)", (clean, sku))
        conn.close()

        upload_status.update({"is_running": False, "message": f"Completado: {upload_status['done']} imágenes."})
    except Exception as e:
        upload_status.update({"is_running": False, "message": f"Error: {e}"})


@router.post("/images/bulk-upload")
async def bulk_upload_images(request: Request, files: list[UploadFile], overwrite: bool = False):
    _require_auth(request)
    if upload_status["is_running"]:
        raise HTTPException(status_code=409, detail="Ya hay una carga en proceso")
    folder_id = _get_folder_id()

    file_data = []
    for file in files:
        stem = Path(file.filename or "").stem
        parts = stem.rsplit("-", 1)
        if len(parts) != 2 or not parts[1].isdigit():
            raise HTTPException(status_code=400, detail=f"{file.filename}: nombre inválido (esperado SKU-N.ext)")
        data = await file.read()
        mime = (file.content_type or "image/jpeg").split(";")[0].strip()
        file_data.append({
            "sku": parts[0].upper(),
            "idx": int(parts[1]),
            "data": data,
            "mime": mime,
            "filename": file.filename or "",
        })

    threading.Thread(target=_run_bulk_upload, args=(file_data, folder_id), daemon=True).start()
    return {"message": f"Procesando {len(file_data)} imágenes en background."}
